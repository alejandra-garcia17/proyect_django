import io
from django.contrib import messages
from django.shortcuts import redirect
from django.http import HttpResponse

# Librerías para la generación de archivos (Excel y PDF)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


# === MIXIN DE SEGURIDAD (Tu mixin original) ===
class StaffRequiredMixin:
    """
    Mixin que verifica si el usuario es miembro del staff.
    Si no es staff, redirige con mensaje de error.
    """
    staff_redirect_url = '/'
    staff_error_message = 'You do not have permission to perform this action. Staff access required.'

    def dispatch(self, request, *args, **kwargs):
        """
        Interceptamos aquí para verificar permisos ANTES de
        procesar la petición (GET o POST).
        """
        if not request.user.is_staff:
            messages.error(request, self.staff_error_message)
            return redirect(self.staff_redirect_url)

        return super().dispatch(request, *args, **kwargs)


# === MIXIN DE EXPORTACIÓN (Nuevo añadido) ===
class ExportFieldsMixin:
    """
    Mixin genérico para Django ListView que permite exportar el QuerySet filtrado
    a formatos Excel y PDF basándose en un parámetro 'export' en la URL.
    """
    export_fields = []       # Campos del modelo (ej: ['name', 'brand__name'])
    export_headers = []      # Cabeceras de las columnas (ej: ['Nombre', 'Marca'])
    export_filename = 'export_data'

    def get_export_fields_and_headers(self):
        """Si no se definen campos específicos, toma todos los campos directos del modelo."""
        if not self.export_fields:
            fields = [field.name for field in self.model._meta.fields]
            return fields, [f.capitalize() for f in fields]
        return self.export_fields, self.export_headers

    def get(self, request, *args, **kwargs):
        export_type = request.GET.get('export')
        
        # Si no se solicita exportar, continúa con el comportamiento normal del ListView HTML
        if export_type not in ['excel', 'pdf']:
            return super().get(request, *args, **kwargs)

        # Obtenemos el queryset con todos los filtros de búsqueda aplicados (sin paginar)
        self.object_list = self.get_queryset()
        fields, headers = self.get_export_fields_and_headers()

        if export_type == 'excel':
            return self.export_to_excel(self.object_list, fields, headers)
        elif export_type == 'pdf':
            return self.export_to_pdf(self.object_list, fields, headers)

    def _get_nested_value(self, obj, field_path):
        """Permite obtener valores de relaciones FK usando sintaxis de doble guion bajo (__)."""
        parts = field_path.split('__')
        val = obj
        for part in parts:
            if val is None:
                return ""
            if part == 'is_active' and hasattr(val, 'is_active'):
                return "Activo" if getattr(val, 'is_active') else "Inactivo"
            val = getattr(val, part, "")
        return val

    def export_to_excel(self, queryset, fields, headers):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for obj in queryset:
            row_data = [self._get_nested_value(obj, field) for field in fields]
            ws.append(row_data)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="{self.export_filename}.xlsx"'
        wb.save(response)
        return response

    def export_to_pdf(self, queryset, fields, headers):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=18, leading=22, textColor=colors.HexColor('#1F4E78'), spaceAfter=12)
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=9, leading=11)
        header_style = ParagraphStyle('HeaderStyle', parent=styles['Normal'], fontSize=10, bold=True, leading=12, textColor=colors.white)

        story.append(Paragraph(f"Reporte de {self.model._meta.verbose_name_plural.capitalize()}", title_style))
        story.append(Spacer(1, 10))

        table_data = [[Paragraph(h, header_style) for h in headers]]
        for obj in queryset:
            row = []
            for field in fields:
                val = str(self._get_nested_value(obj, field))
                row.append(Paragraph(val, cell_style))
            table_data.append(row)

        table = Table(table_data, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F2F2F2'), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D3D3D3')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))

        story.append(table)
        doc.build(story)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.export_filename}.pdf"'
        return response